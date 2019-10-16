import barcode
from barcode.writer import ImageWriter
import requests
import csv
from tkinter import filedialog
import os
import openpyxl as op
import tkinter as tk
import sys
root = tk.Tk()
root.withdraw()


current_location = '262'


def barcode_build(employee_name, employee_code, barcode_folder):
    ''' this created a full barcode and safe it as .png , it takes the name and the code'''
    try:
        EAN = barcode.get_barcode_class('code128')
        ean = EAN(u'{}'.format(employee_code), writer=ImageWriter())
        ean.save('{}\{}'.format(barcode_folder, employee_name))
    except:
        pass


def getting_codes():

    key_api = 'spbdxdqepugjhgeh4kbrn6t2sz7jfytm'

    def get_data(out_file, code=37):
        #  this is preconfigure for current employee status  for the whole company
        #  by we can use any other code that not require a date input.
        ship_api_url = "https://api.mytimestation.com/v0.1/reports/?api_key={}&id={}&exportformat=csv".format(key_api,
                                                                                                              code)
        url_responds = requests.get(ship_api_url)

        data = url_responds.text
        save_text = open(out_file, 'w', encoding='utf8')
        save_text.write(data)
        save_text.close()


    def processing_data():

        out_file = "employee_codes.csv"
        get_data(out_file, code=35)

        dict_codes = {}
        with open(out_file, 'r', newline="", encoding='utf8') as f:
            reader = csv.DictReader(f, delimiter=',')

            for row in reader:
                dict_codes[row['Name']] = (row['QR Code'], row['Card Number'])

        return dict_codes

    return processing_data()


def sheet_info_loader():
    try:
        path_source = os.path.normpath(filedialog.askopenfilename(title = 'Please introduce the sheet with names',
                                                                  filetypes=[('Excel File', '*.xlsx')]))
        load_sheet = op.load_workbook(path_source)
        main_sheet = load_sheet.sheetnames[0]
        working_sheet = load_sheet[main_sheet]

        mylist = []
        for row in working_sheet['A{}:A{}'.format(working_sheet.min_row, working_sheet.max_row)]:
            for cell in row:
                mylist.append(cell.value)
        return mylist

    except:

        mylist = []
        items = sys.stdin.readlines()
        print(items)



def compare_list():

    filter_employees = {}
    server_list = getting_codes()
    employees_list = sheet_info_loader()
    for names in employees_list:
        if names in server_list:
            filter_employees[names] = server_list[names]

    return filter_employees


def created_barcodes():

    list_employees = compare_list()
    path_save = os.path.normpath(filedialog.askdirectory(title='please choose folder where is going to be save'))

    for item in list_employees:
        barcode_build(item, list_employees[item][0], path_save)


if __name__ == '__main__':

    created_barcodes()
